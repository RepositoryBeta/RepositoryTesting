from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox, Treeview
import os
import hashlib
import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore
from openpyxl.styles import Font, Alignment, PatternFill
from collections import Counter

'''CODIFICACIÓN DE SISTEMA DE ANOTADO'''

class Base_firebase():
    def __init__(self):
        self.cred = credentials.Certificate({
  "type": "service_account",
  "project_id": "corpusderecho",
  "private_key_id": "bd2d86228c1074e220dd1195e91fbfeb45c8ac0e",
  "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQDF4rcieLaBdoy5\nMr9+IfiKB+SnjNv4hTEmBB2PimMNSa0moLgp7jjnj1mWl2zO5BBDrF3LwOPTZEHQ\nDOAc2O5gqmkKFPsz+9dC2+xf17tR7SYmfHj6Ge+9zJyrnrTMPhT3Q6UClazDD8JF\nDQqZYzYgVpT6285LD+M5vRsCHJSOq84ZBONJ+TwzyOlkfSYbMEurtlohdZ1IFxUT\nPcl2DGcJb+rHVJT0ESQUh1yNPXRtsMCgU87+n8vevKkeszonGe1K2qsdXbIWGtB8\nBXOfKicGD7EC7CFrB5RLGH+tiCfyv66dBmrSA39UFrc/3/QNQWxCH3wAv/8m8QJn\n9n/MfCAHAgMBAAECggEAWELx6gcrZ/0ncScocCrYv0AcBXt3VZaawFkPkkloOrIE\n7/i5i8I8UVQH4noxxunEM1r7Dvk+5LeP/yru/w7m6+i2JxAKpzoJ3kFXYbxISZdb\n78iO8QaGfwJ3Urjb7Uk3cr4SKF/4CZzKyp+xfY7Zwny7wFQv5Bz2/XbSSBPVUviW\ndE0ZRdjHbUOKBOXZuH+weSnDMr5uO2wAD9/ePk+O6Ahzc2Oaz5vsgf8Rb52xbL2L\nNPoBuiFf7zKAgb2iv0ckd2/HiUAfIIbje1xKid5+IumQMnKJyt3OnetHC95emryS\nN8ywX7fFDHJxj68oujpRRsQ+oODa4lNXy2PmdDPl4QKBgQD6ExfsJ1hl3VUZzsHd\n4OtGdnLPlSLV2zh8oxKVx8PzUAhlwxdYN+RQrf641i6nNeHJ6Ivc1M+UR6HETVOo\n6zT5EnRrw3A8WVPnfxoWdmQwIpfNv4B9u0slpQHAOnsCGDNFC1uAdHVLjbNdm5ml\nQD6aPbQKrjzdt1taQZYXvwp9MQKBgQDKkw2cCW56NHsn/caTBm/s2FdbZD32PLVM\nTrbn2p3q8rZjagfq+Fo5rxHcsj1pEYc81BfdG7rY9PnzEJGLQMy5K4O1gvhBR/cI\nxtVCkbEYLqgo8MZye0EBAxAXmrq7iqG88vNMIl5r7aCp43b9LKwfS6s2tIIQcbF4\ne8AjvZZCtwKBgQDO4UbUJgSaWM2TSRnW/cCGDW2Tz1P/SZlket/gK2AvBNG9pfTF\nFY/7MsZckUwNJqVBDPHP4fpMOCpapIIGZ3buatcHY6qRMnEYPW0Okij2LVgabUfW\nPEDBxs05AuhKZkzb3LZ5AV8b/wEY+qXUMN7ZaRxuDZDq90pK06FgybnasQKBgD9N\ntPMN/3gw7kuRAnzvhgp9kM7+hF5umjk+X+oUZ0UEk/p9J/xFn+xmyS4uhQJQd/KK\nVOrDv6AM7g8HHnRly5QSBRaz1uCPnKMhf4NOBAmKiwJAO8OKwMWQL3cw4ym4KNLj\nGSSxmNAhchOlAYoca61b+CUgQ4TgApWkFmcU1+ABAoGAM8OiqUetSK43Ys7bzZLg\nGaa1Cw5o1pJFFiPRvRrzs59ZkA9/MrRN2T7PywFLniv2his+IrNU3BwrV9epFtjc\n3QoM36eb5ttoYMr6T+TDWVnK+ERlH5N8r5VZAgFRNWlPamRTMkpjPSrjvs/qCMKf\nN1TReOOyrmYCHhxNOhG67G4=\n-----END PRIVATE KEY-----\n",
  "client_email": "firebase-adminsdk-9r39w@corpusderecho.iam.gserviceaccount.com",
  "client_id": "113929231003963747832",
  "auth_uri": "https://accounts.google.com/o/oauth2/auth",
  "token_uri": "https://oauth2.googleapis.com/token",
  "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
  "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-9r39w%40corpusderecho.iam.gserviceaccount.com"
})
        if not firebase_admin._apps:
            firebase_admin.initialize_app(self.cred)
            self.conexion = self.Conexion()
        else:
            self.conexion = self.Conexion()

    def Conexion(self):
        return firestore.client()

    def Registrar(self,usu_reg):
        self.conexion.collection('usuarios').add(usu_reg)

    def Logear(self,usu,passw):
        return self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).where(u'contrasenia',u'==',passw)

    def Conf_regis(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu)
        return [doc.id for doc in result.stream()]

    def Conf_Ced(self,ced):
        result = self.conexion.collection(u'usuarios').where(u'cedula',u'==',ced)
        return [doc.id for doc in result.stream()]

    def Conf_Seleccion(self,sem):
        result = self.conexion.collection(u'usuarios').where(u'semestre',u'==',sem).stream()
        cuentaA = 0
        cuentaB = 0
        cuentaC = 0
        seleccion = ""
        for doc in result:
            result2 = doc.to_dict()
            try: 
                with switch(result2["seleccion"], comparator=lambda x, y: x in y) as s:
                    if s.case(("Seleccion A"), True):
                        cuentaA = cuentaA + 1
                    if s.case(("Seleccion B"), True):
                        cuentaB = cuentaB + 1
                    if s.case(("Seleccion C"), True):
                        cuentaC = cuentaC + 1
            except KeyError:
                seleccion = "Seleccion A"

        if (cuentaA < 3):
            seleccion = "Seleccion A"
        elif (cuentaA == 3 and cuentaB < 3):
            seleccion = "Seleccion B"
        elif (cuentaA == 3 and cuentaB == 3 and cuentaC < 3):
            seleccion = "Seleccion C"
        return seleccion

    def Obt_nombres(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["nombres"])

    def Obt_apellidos(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["apellidos"])

    def Obt_carrera(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["carrera"])

    def Obt_semestre(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["semestre"])
        
    def Obt_nivel(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["nivel"])

    def Obt_seleccion(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return str(result2["seleccion"])

    def Obt_sem_niv_sel(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return result2["semestre"] + " - " + result2["nivel"] + " - " + result2["seleccion"]

    def Obt_ruta_textos(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu).stream()
        for doc in result:
            result2 = doc.to_dict()
        return "./TRANSCRIPCIONES/"+ result2["carrera"] + "/" + result2["semestre"] + "/" + result2["nivel"] + "/" + result2["seleccion"] + "/"

    def Add_semestres_textos(self,id_usu,usu,carr,sem,niv,sel):
        ref_usuario = self.conexion.collection('usuarios').document(id_usu[0])
        ref_carrera = ref_usuario.collection('carrera').document(carr)
        ref_semestre = ref_carrera.collection('semestres').document(sem)
        ref_nivel = ref_semestre.collection('nivel').document(niv)
        ref_seleccion = ref_nivel.collection('seleccion').document(sel)
        lista_textos = os.listdir(self.Obt_ruta_textos(usu))
        for lis_t in lista_textos:
            data = {
                'texto': lis_t,
                'estado': 0
            }
            ref_seleccion.collection('textos').add(data)    

    def lista_textos_base(self, usuario):
        ref_usuario = self.conexion.collection('usuarios').document(usuario)
        ref_carrera = ref_usuario.collection('carrera').document(self.Obt_carrera(usuario))
        ref_semestre = ref_carrera.collection('semestres').document(self.Obt_semestre(usuario))
        ref_nivel = ref_semestre.collection('nivel').document(self.Obt_nivel(usuario))
        ref_seleccion = ref_nivel.collection('seleccion').document(self.Obt_seleccion(usuario))
        result = ref_seleccion.collection('textos').where('estado','==',0)
        return [doc.to_dict() for doc in result.stream()]

    def Obtener_id_texto(self,usuario,texto):
        ref_usuario = self.conexion.collection('usuarios').document(usuario)
        ref_carrera = ref_usuario.collection('carrera').document(self.Obt_carrera(usuario))
        ref_semestre = ref_carrera.collection('semestres').document(self.Obt_semestre(usuario))
        ref_nivel = ref_semestre.collection('nivel').document(self.Obt_nivel(usuario))
        ref_seleccion = ref_nivel.collection('seleccion').document(self.Obt_seleccion(usuario))
        result = ref_seleccion.collection('textos')
        return [doc.to_dict() for doc in result.stream()]

    def Fin_anotado_text(self,usuario,texto):
        id_texto = self.Obtener_id_texto(usuario,texto)
        ref_usuario = self.conexion.collection('usuarios').document(usuario)
        ref_carrera = ref_usuario.collection('carrera').document(self.Obt_carrera(usuario))
        ref_semestre = ref_carrera.collection('semestres').document(self.Obt_semestre(usuario))
        ref_nivel = ref_semestre.collection('nivel').document(self.Obt_nivel(usuario))
        ref_seleccion = ref_nivel.collection('seleccion').document(self.Obt_seleccion(usuario))
        result = ref_seleccion.collection('textos').document(id_texto[0])
        result.update({'estado': 1})

    def Obtener_Cedula(self,usu):
        result = self.conexion.collection(u'usuarios').where(u'usuario',u'==',usu)
        return [doc.to_dict() for doc in result.stream()]

    def Consultar_seleccion(self,seleccion):
        result = self.conexion.collection('anotado').where('Palabra','==',seleccion['Palabra']).\
            where('Cedula','==',seleccion['Cedula']).where('Nombre_archivo','==',seleccion['Nombre_archivo'])
        return [doc.id for doc in result.stream()]

    def Guardar_seleccion(self,seleccion):
        self.conexion.collection('anotado').add(seleccion)

    def Consultar_anotado(self):
        estudiantes = self.Consultar_estudiante()
        resultAnotado = self.conexion.collection('anotado')
        anotado = [doc.to_dict() for doc in resultAnotado.stream()]
        for r in anotado:
            estudiante = next((x for x in estudiantes if x['cedula'] == r['Cedula']), None)
            if estudiante is not None:
                r.update({'Nivel': estudiante['nivel']})
                r.update({'Seleccion': estudiante['seleccion']})
            else:
                r.update({'Nivel': ''})
                r.update({'Seleccion': ''})
        return anotado

    def Consultar_estudiante(self):
        result = self.conexion.collection('usuarios')
        return [doc.to_dict() for doc in result.stream()]


class Login(Frame):

    def __init__(self,master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Login")
        self.master.resizable(False,False)
        self.pack()
        self.crear_widget()
        self.base = Base_firebase()
        self.master.protocol("WM_DELETE_WINDOW", self.master.quit)

    # LOGEO AL SISTEMA#
    def logear(self):
        self.usuario= self.usuario_txt_log.get()
        self.password = hashlib.md5(self.user_pass_txt_log.get().encode()).hexdigest()
        result = self.base.Logear(self.usuario,self.password)
        result2 = [doc.to_dict() for doc in result.stream()]
        if(len(result2)>0):
            self.id_usuario = self.base.Conf_regis(self.usuario)
            for r in result2:
                if(r['rol']== 1):
                    self.Vent_Anotado()
                else:
                    self.Vent_Admin()
        else:
            messagebox.showwarning("¡Atención!", "Usuario y/o Clave incorrecta")

    def Vent_Admin(self):
        self.master.withdraw()
        self.inst_admin = Toplevel()
        self.vent_admin = Admin(self.base,self.usuario,self.id_usuario,self.inst_admin)

    def Vent_Anotado(self):
        self.master.withdraw()
        self.inst_anotado = Toplevel()
        self.vent_anotado = Anotado(self.base,self.usuario,self.id_usuario,self.inst_anotado)

    def Vent_Registro(self):
        self.master.withdraw()
        self.inst_registro = Toplevel()
        self.vent_registro = Registro(self.base,self.inst_registro)

    def crear_widget(self):
        self.miFrame = Frame(self)
        self.miFrame.config(bg="white")
        self.miFrame.pack()
        self.miFrame.config(cursor="hand2")

        # --------------MENSAJE DE BIENVENIDA-----------------#
        self.lb_softsis_log = Label(self.miFrame, text="CEDUG")
        self.lb_softsis_log.grid(row=0, column=0, padx=10, pady=5, columnspan=2)
        self.lb_softsis_log.config(bg="white", fg="gray",font=("Fixedsys",24,"bold"))
        self.lb_softsis_log = Label(self.miFrame, text="(Corpus Etiquetado de la Carrera de Derecho de la Universidad de Guayaquil)")
        self.lb_softsis_log.grid(row=1, column=0, padx=10, pady=3, columnspan=2)
        self.lb_softsis_log.config(bg="white", fg="gray",font=("Fixedsys",24,"bold"))
        self.lb_0 = Label(self.miFrame, text="Sistema de Anotación y Etiquetado de las Palabras Complejas")
        self.lb_0.grid(row=2, column=0, padx=10, pady=2, columnspan=2)
        self.lb_0.config(bg="white", fg="gray",font=("Fixedsys",24,"bold"))
        self.lb_1 = Label(self.miFrame, text="Carrera de Derecho")
        self.lb_1.grid(row=3, column=0, padx=10, pady=2, columnspan=2)
        self.lb_1.config(bg="white", fg="gray",font=("Fixedsys",13,"bold"))
        self.lb_2 = Label(self.miFrame, text="Facultad de Jurisprudencia, Ciencias Sociales y Políticas ")
        self.lb_2.grid(row=4, column=0, columnspan=2)
        self.lb_2.config(bg="white", fg="gray",font=("Fixedsys",13,"bold"))
        self.lb_2 = Label(self.miFrame, text="Universidad de Guayaquil")
        self.lb_2.grid(row=5, column=0, columnspan=2)
        self.lb_2.config(bg="white", fg="gray", font=("Fixedsys", 13, "bold"))

        # --------------USUARIO-----------------#
        self.lb_user_log = Label(self.miFrame, text="Usuario")
        self.lb_user_log.grid(row=6, column=0, padx=10, pady=15, columnspan=2)
        self.lb_user_log.config(bg="white", fg="gray",font=("Fixedsys",13,"bold"))
        self.usuario_txt_log = Entry(self.miFrame)
        self.usuario_txt_log.grid(row=7, column=0, padx=10, columnspan=2)
        self.usuario_txt_log.config(font=("Fixedsys",12,"bold"), fg="blue")

        # --------------CONTRASEÑA-----------------#
        self.lb_pass_log = Label(self.miFrame, text="Password")
        self.lb_pass_log.grid(row=8, column=0, padx=10, pady=15, columnspan=2)
        self.lb_pass_log.config(bg="white", fg="gray",font=("Fixedsys",13,"bold"))
        self.user_pass_txt_log = Entry(self.miFrame)
        self.user_pass_txt_log.grid(row=9, column=0, padx=10, columnspan=2)
        self.user_pass_txt_log.config(show="*",font=("Fixedsys",12,"bold"), fg="blue")

        # ---------------Botones-------------------#

        '''               #GUARDAR#                 '''
        self.btn_ingresar = Button(self.miFrame, text="Ingresar", width=10, command=self.logear)
        self.btn_ingresar.grid(row=10, column=0, padx=10, pady=20)
        self.btn_ingresar.config(bg="white", fg="gray",font=("Fixedsys",12,"bold"))

        '''               #REGISTRAR#               '''
        self.btn_registrar = Button(self.miFrame, text="Registrar", width=10, command=self.Vent_Registro)
        self.btn_registrar.grid(row=10, column=1, padx=10, pady=20)
        self.btn_registrar.config(bg="white", fg="gray",font=("Fixedsys",12,"bold"))


class Admin(Frame):

    def __init__(self,bd,usu,id_usu,master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Visualización de Anotado")
        self.master.resizable(False,False)
        self.pack()
        self.base = bd
        self.usuario = usu
        self.id_usuario = id_usu
        self.crear_widget()
        self.master.protocol("WM_DELETE_WINDOW", self.master.quit)

    def crear_excel(self):
        s_font = Font(name='Calibri',size=11,bold=True,italic=False,underline='none',strike=False,color='000000')
        s_alignment = Alignment(horizontal='center', vertical='center')
        s_pattern = PatternFill(fill_type='solid', fgColor='00B0F0')
        style_2 = Font(color="000000", italic=False, size=11)
        self.excel = openpyxl.Workbook()
        hoja_anotado = self.excel.active
        hoja_anotado.title = "Datos del Anotado"
        self.excel.create_sheet("Datos de estudiantes")
        hoja = self.excel.active
        anotado = self.base.Consultar_anotado()
        estudiante = self.base.Consultar_estudiante()
        hoja.append(('N°', 'Palabra', 'Cedula', 'Nombres', 'Apellidos', 'Nombre_archivo', 'Posicion_en_el_texto', 'Longitud', 'Repeticiones'))
        cont = 1
        for cell in hoja["1:1"]:
            cell.font = s_font
            cell.alignment = s_alignment
            cell.fill = s_pattern
            hoja.column_dimensions['A'].width = float(14.50)
            hoja.column_dimensions['B'].width = float(14.50)
            hoja.column_dimensions['C'].width = float(14.50)
            hoja.column_dimensions['D'].width = float(87.50)
            hoja.column_dimensions['E'].width = float(20.00)
            hoja.column_dimensions['F'].width = float(14.00)
            hoja.column_dimensions['G'].width = float(14.00)
            hoja.column_dimensions['H'].width = float(14.00)
            hoja.column_dimensions['I'].width = float(14.00)
        if(len(anotado) > 0):
            for a in anotado:
                hoja.append(
                    (
                        cont,
                        a['Palabra'],
                        a['Cedula'],
                        a['Nombres'],
                        a['Apellidos'],
                        a['Nombre_archivo'],
                        a['Posicion_en_el_texto'],
                        a['Longitud'],
                        a['Repeticiones']
                    )
                )
                cont += 1
            for row in hoja.iter_rows(min_row=2, max_col=9, max_row=cont+2):
                for cell in row:
                    cell.font = style_2
                    cell.alignment = s_alignment
        hoja = self.excel['Datos de estudiantes']
        hoja.append(('N°', 'Cedula', 'Nombres', 'Apellidos', 'Edad', 'Carrera', 'Semestre', 'Promedio', 'Nivel', 'Seleccion'))
        cont = 1
        for cell in hoja["1:1"]:
            cell.font = s_font
            cell.alignment = s_alignment
            cell.fill = s_pattern
            hoja.column_dimensions['A'].width = float(14.50)
            hoja.column_dimensions['B'].width = float(14.50)
            hoja.column_dimensions['C'].width = float(50.00)
            hoja.column_dimensions['D'].width = float(50.00)
            hoja.column_dimensions['E'].width = float(23.00)
            hoja.column_dimensions['F'].width = float(14.00)
            hoja.column_dimensions['G'].width = float(14.00)
            hoja.column_dimensions['H'].width = float(14.00)
            hoja.column_dimensions['I'].width = float(14.00)
            hoja.column_dimensions['J'].width = float(14.00)
        if(len(estudiante) > 0):
            for e in estudiante:
                hoja.append(
                    (
                        cont,
                        e['cedula'],
                        e['nombres'],
                        e['apellidos'],
                        e['edad'],
                        e['carrera'],
                        e['semestre'],
                        e['promedio'],
                        e['nivel'],
                        e['seleccion']
                    )
                )
                cont += 1
            for row in hoja.iter_rows(min_row=2, max_col=8, max_row=cont+2):
                for cell in row:
                    cell.font = style_2
                    cell.alignment = s_alignment
        self.descargar_excel()

    def descargar_excel(self):
        self.excel.save('Datos del Anotado.xlsx')
        messagebox.showinfo("¡Descargado!", "Archivo en excel descargado")

    def mostrar_datos(self):
        if(self.estudiante == False and self.anotado == False):
            messagebox.showwarning("¡Atención!", "Elija una tabla")
        else:
            if(self.anotado == True):
                self.mostrar_datos_anotado()
            else:
                self.mostrar_datos_estudiante()

    def mostrar_datos_anotado(self):
        self.anotado = True
        self.estudiante = False
        self.tabla.delete(*self.tabla.get_children())
        self.tabla.column("#0", width=40, minwidth=40,anchor="center")
        self.tabla.column("#1", width=150, minwidth=150,anchor="center")
        self.tabla.column("#2", width=80, minwidth=80,anchor="center")
        self.tabla.column("#3", width=80, minwidth=80,anchor="center")
        self.tabla.column("#4", width=80, minwidth=80,anchor="center")
        self.tabla.column("#5", width=600, minwidth=600,anchor="center")
        self.tabla.column("#6", width=80, minwidth=80,anchor="center")
        self.tabla.column("#7", width=80, minwidth=80,anchor="center")
        self.tabla.column("#8", width=80, minwidth=80,anchor="center")
        self.tabla.heading("#0",text='N°', anchor="center")
        self.tabla.heading("#1",text="Palabra", anchor="center")
        self.tabla.heading("#2",text="Cedula", anchor="center")
        self.tabla.heading("#3",text="Nombres", anchor="center")
        self.tabla.heading("#4",text="Apellidos", anchor="center")
        self.tabla.heading("#5",text="Nombre_archivo", anchor="center")
        self.tabla.heading("#6",text="Posicion", anchor="center")
        self.tabla.heading("#7",text="Longitud", anchor="center")
        self.tabla.heading("#8",text="Repeticiones", anchor="center")
        self.tabla.heading("#9",text="Nivel", anchor="center")
        self.tabla.heading("#10",text="Selección", anchor="center")
        cont=1
        result = self.base.Consultar_anotado()
        if(len(result) > 0):
            for r in result:
                self.tabla.insert("",cont,text=str(cont),values=(r['Palabra'],r['Cedula'],r['Nombres'],r['Apellidos'],r['Nombre_archivo'],r['Posicion_en_el_texto'],
                                                                 r['Longitud'],r['Repeticiones'], r['Nivel'], r['Seleccion']))
                cont+=1
        else:
            messagebox.showwarning("¡Atención!", "No hay datos")

    def mostrar_datos_estudiante(self):
        self.estudiante = True
        self.anotado = False
        self.tabla.delete(*self.tabla.get_children())
        self.tabla.column("#0", width=50, minwidth=50,anchor="center")
        self.tabla.column("#1", width=120, minwidth=120,anchor="center")
        self.tabla.column("#2", width=250, minwidth=250,anchor="center")
        self.tabla.column("#3", width=250, minwidth=250,anchor="center")
        self.tabla.column("#4", width=110, minwidth=110,anchor="center")
        self.tabla.column("#5", width=200, minwidth=200,anchor="center")
        self.tabla.column("#6", width=110, minwidth=110,anchor="center")
        self.tabla.column("#7", width=110, minwidth=110,anchor="center")
        self.tabla.column("#8", width=110, minwidth=110,anchor="center")
        self.tabla.column("#9", width=110, minwidth=110,anchor="center")
        self.tabla.heading("#0",text='N°', anchor="center")
        self.tabla.heading("#1",text="Cedula", anchor="center")
        self.tabla.heading("#2",text="Nombres", anchor="center")
        self.tabla.heading("#3",text="Apellidos", anchor="center")
        self.tabla.heading("#4",text="Edad", anchor="center")
        self.tabla.heading("#5",text="Carrera", anchor="center")
        self.tabla.heading("#6",text="Semestre", anchor="center")
        self.tabla.heading("#7",text="Promedio", anchor="center")
        self.tabla.heading("#8",text="Nivel", anchor="center")
        self.tabla.heading("#9",text="Seleccion", anchor="center")
        cont=1
        result = self.base.Consultar_estudiante()
        if(len(result)):
            for r in result:
                self.tabla.insert("",cont,text=str(cont),values=(r['cedula'],r['nombres'],r['apellidos'],r['edad'],r['carrera'],
                                                                 r['semestre'],r['promedio'],r['nivel'],r['seleccion']))
                cont+=1
        else:
            messagebox.showwarning("¡Atención!", "No hay datos")

    def crear_widget(self):

        #---------------------FRAME DE BOTONES----------------------#
        self.inicio_e = 1
        self.inicio_a = 1
        self.anotado = False
        self.estudiante = False
        self.Frame_tab = Frame(self)
        self.Frame_tab.config(height=35, cursor="hand2", relief="groove")
        self.Frame_tab.grid(row=0, column=0)
        self.btn_act = Button(self.Frame_tab, text="Actualizar", width=23, command=self.mostrar_datos)
        self.btn_act.grid(row=0, column=0, sticky="w")
        self.btn_act.config(bg="gray", fg="white")
        self.lb = Label(self.Frame_tab, text="", width=14)
        self.lb.grid(row=0, column=1, sticky="w")
        self.lb1 = Label(self.Frame_tab, text="", width=14)
        self.lb1.grid(row=0, column=2, sticky="w")
        self.btn_descargar = Button(self.Frame_tab, text="Descargar", width=23, command=self.crear_excel)
        self.btn_descargar.grid(row=0, column=3, sticky="w")
        self.btn_descargar.config(bg="gray", fg="white")
        self.lb2 = Label(self.Frame_tab, text="", width=14)
        self.lb2.grid(row=0, column=4, sticky="w")
        self.lb3 = Label(self.Frame_tab, text="", width=14)
        self.lb3.grid(row=0, column=5, sticky="w")
        self.btn_cerrar = Button(self.Frame_tab, text="Cerrar Sesión", width=23, command=self.master.quit)
        self.btn_cerrar.grid(row=0, column=6, sticky="w")
        self.btn_cerrar.config(bg="gray", fg="white")

        #---------------------FRAME DE TABLA----------------------#
        self.Frame_tabla = Frame(self)
        self.Frame_tabla.config(cursor="hand2", relief="groove", width=20)
        self.Frame_tabla.grid(row=1, column=0)
        self.tabla = Treeview(self.Frame_tabla,height=20,columns = ('#1','#2','#3','#4','#5','#6','#7','#8','#9','#10'))
        self.tabla.grid(row=0,column=0,columnspan=7)
        self.barra = Scrollbar(self.Frame_tabla,command=self.tabla.yview)
        self.barra.grid(row=0, column=7, sticky="nse")
        self.tabla.config(yscrollcommand=self.barra.set)
        self.tabla.column("#0", width=40, minwidth=40,anchor="center")
        self.tabla.column("#1", width=150, minwidth=150,anchor="center")
        self.tabla.column("#2", width=80, minwidth=80,anchor="center")
        self.tabla.column("#3", width=600, minwidth=600,anchor="center")
        self.tabla.column("#4", width=80, minwidth=80,anchor="center")
        self.tabla.column("#5", width=80, minwidth=80,anchor="center")
        self.tabla.column("#6", width=80, minwidth=80,anchor="center")
        self.tabla.column("#7", width=80, minwidth=80,anchor="center")
        self.tabla.column("#8", width=80, minwidth=80,anchor="center")
        self.tabla.column("#9", width=80, minwidth=80,anchor="center")

        #---------------------FRAME DE ESTUDIANTE/ANOTADO----------------------#
        self.Frame_btn = Frame(self)
        self.Frame_btn.config(cursor="hand2", relief="groove")
        self.Frame_btn.grid(row=2, column=0)
        self.btn_est = Button(self.Frame_btn, text="Estudiante", width=23, command=self.mostrar_datos_estudiante)
        self.btn_est.grid(row=0, column=0, sticky="w")
        self.btn_est.config(bg="gray", fg="white")
        self.btn_anot = Button(self.Frame_btn, text="Anotado", width=23, command=self.mostrar_datos_anotado)
        self.btn_anot.grid(row=0, column=1, sticky="w",)
        self.btn_anot.config(bg="gray", fg="white")


class Registro(Frame):

    def __init__(self,bd,master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Formulario de Registro")
        self.master.resizable(False,False)
        self.pack()
        self.base = bd
        self.crear_widget()
        self.master.protocol("WM_DELETE_WINDOW", self.master.quit)

    def usuario_duplicado(self, usu):
        result_usu = self.base.Conf_regis(usu)
        if (len(result_usu) > 0):
            return True
        else:
            return False

    def usuario_diferente(self, usu):
        result_usu = self.base.Conf_regis(usu)
        if (len(result_usu) > 0):
            usua = usu + str(len(result_usu))
        return usua

    def validar(self):
        validacion = 0
        val_ced = 0
        coeficientes = [2, 1, 2, 1, 2, 1, 2, 1, 2]
        decenas = [10, 20, 30, 40, 50, 60, 70, 80, 90]
        suma = 0
        res_ult = 0
        if(self.cedula_txt.get()!="" and self.nombre_txt.get()!="" and self.apellidos_txt.get()!="" and self.edad_txt.get()!="" and
                self.carrera_txt.get()!="" and self.semestre_txt.get()!=""): #and self.promedio_txt.get()!=""#
            ced = list(map(int, str(self.cedula_txt.get())))
            dos_dig = int(str(ced[0]) + str(ced[1]))
            tercer_dig = int(str(ced[2]))
            decimo_dig = int(str(ced[9]))
            if (len(ced) < 11):
                if (dos_dig > 0 and dos_dig < 25):
                    if (tercer_dig >= 0 and tercer_dig < 6):
                        for c in range(len(ced) - 1):
                            result = ced[c] * coeficientes[c]
                            if (result >= 10):
                                result -= 9
                            suma += result
                        while True:
                            for d in decenas:
                                if (suma <= d):
                                    res_ult = d - suma
                                    break
                            break
                        if (res_ult == decimo_dig):
                            self.cedula = self.cedula_txt.get()
                        else:
                            validacion = 1
                            val_ced = 1
                    else:
                        validacion = 1
                        val_ced = 1
                else:
                    validacion = 1
                    val_ced = 1
            else:
                validacion = 1
                val_ced = 1
                messagebox.showwarning("Error", "La cedula no puede exceder de 10 dígitos")
            if(val_ced == 0):
                self.cedula = self.cedula_txt.get()
            else:
                messagebox.showwarning("Error", "Cedula incorrecta")
            if(len(self.nombre_txt.get()) > 5 and len(self.apellidos_txt.get()) > 5 ):
                self.nombre = self.nombre_txt.get().lower()
                self.apellidos = self.apellidos_txt.get().lower()
                self.edad = self.edad_txt.get()
                p_letra = list(map(str, self.nombre))
                pri_letra = p_letra[0]
                p_apellido = self.apellidos.split()
                self.usu = ""
                self.contra = ""
                try:
                    pri_apellido = p_apellido[0]
                    seg_apellido = p_apellido[1]
                    pri_seg_apellido = list(map(str, seg_apellido))
                    self.usu = pri_letra + pri_apellido + pri_seg_apellido[0]
                    u = 1
                    while self.usuario_duplicado(self.usu):
                        try:
                            self.usu += pri_seg_apellido[u]
                            u += 1
                        except IndexError:
                            self.usu = self.usuario_diferente(self.usu)
                            break
                    self.contra = self.usu + "123"
                    self.contrasenia = hashlib.md5(self.contra.encode()).hexdigest()
                    self.usuario = self.usu
                except IndexError:
                    validacion = 1
                    messagebox.showwarning("Error", "Debe colocar sus dos apellidos")
                result_ced = self.base.Conf_Ced(self.cedula_txt.get())
                if (len(result_ced) > 0):
                    validacion = 1
                    messagebox.showwarning("Error","La cedula que trata de registrar ya existe")
                else:
                    self.cedula = self.cedula_txt.get()
            else:
                validacion = 1
                messagebox.showwarning("Error", "Los campos de nombres, apellidos, usuario y password deben tener una longitud "
                                                "mayor a 5 caracteres")
            self.carrera = self.carrera_txt.get()
            self.semestre = self.semestre_txt.get()
            self.nivel = self.nivel_txt.get()
            self.seleccion = self.seleccion_txt.get()
            band = 1
            #try:
             #   self.promedio = float(self.promedio_txt.get())
            #except ValueError:
                #band=0
            #if(band==1):
             #   if(self.promedio <= 0.00):
                    #validacion = 1
              #      messagebox.showwarning("Error", "Por favor, ingrese un numero mayor a 0.00")
            #else:
                #validacion = 1
             #   messagebox.showwarning("Error", "Debe ingresar un numero flotante. Ej: 8.00")
        else:
            validacion = 1
            messagebox.showwarning("Error", "Todos los campos deben ser llenados")

        if(validacion==0):
            self.guardar()

    def guardar(self):
        data = {
            'cedula': self.cedula,
            'nombres': self.nombre,
            'apellidos': self.apellidos,
            'edad': self.edad,
            'carrera': self.carrera,
            'semestre': self.semestre,
            'nivel': self.nivel,
            'seleccion': self.seleccion,
#            'promedio': self.promedio,
            'usuario': self.usuario,
            'contrasenia': self.contrasenia,
            'rol': 1
        }
        self.base.Registrar(data)
        self.id_usuario = self.base.Conf_regis(self.usuario)
        if(len(self.id_usuario)>0):
            self.base.Add_semestres_textos(self.id_usuario, self.usuario, self.carrera, self.semestre, self.nivel, self.seleccion)
            messagebox.showinfo("Registro", "El usuario ha sido registrado con éxito"+ "\n" +"Usuario: "+self.usuario+"\n"
                                +"Contraseña: "+self.contra)
            self.master.withdraw()
            self.inst_login = Toplevel()
            self.vent_login = Login(self.inst_login)
        else:
            messagebox.showwarning("Error", "No ha sido posible registrar el usuario")

    def opcionCarrera(self,event):
        op_carrera = self.carrera_txt.current()
        self.semestre_txt.set("")
        if (op_carrera == 0):
            self.semestre_txt["values"] = ["1", "2", "3", "4", "5", "6", "7", "8", "9"]

    def opcionSemestre(self,event):
        nivel = ""
        with switch(self.semestre_txt.get(), comparator=lambda x, y: x in y) as s:
            if s.case(("1", "2", "3"), True):
                nivel = "Basico"
            if s.case(("4", "5", "6"), True):
                nivel = "Intermedio"
            if s.case(("7", "8", "9"), True):
                nivel = "Avanzado"
        self.nivel_txt.set("")
        self.nivel_txt["values"] = [nivel]
        self.nivel_txt.current(0)

        self.seleccion_txt.set("")
        self.seleccion_txt["values"] = [self.base.Conf_Seleccion(self.semestre_txt.get())]
        self.seleccion_txt.current(0)

    def crear_widget(self):
        self.Frame_reg = Frame(self)
        self.Frame_reg.pack()
        self.Frame_reg.config(cursor="hand2")

        # --------------FORMULARIO-----------------#
        self.lb_form = Label(self.Frame_reg, text="Formulario")
        self.lb_form.grid(row=0, column=0, columnspan=2, padx=10, pady=5)
        self.lb_form.config(font=("Fixedsys",24,"bold"))

        # --------------CEDULA-----------------#
        self.lb_cedula = Label(self.Frame_reg, text="Cedula:")
        self.lb_cedula.grid(row=1, column=0, padx=10, pady=10)
        self.lb_cedula.config(font=("Fixedsys",12,"bold"), fg="blue")
        self.cedula_txt = Entry(self.Frame_reg,width=25)
        self.cedula_txt.grid(row=1, column=1, sticky="w", padx=10)
        self.cedula_txt.config(font=("Fixedsys",8,"bold"))

        # --------------NOMBRES-----------------#
        self.lb_nombre = Label(self.Frame_reg, text="Nombres:")
        self.lb_nombre.grid(row=2, column=0, padx=10, pady=10)
        self.lb_nombre.config(font=("Fixedsys",12,"bold"), fg="blue")
        self.nombre_txt = Entry(self.Frame_reg,width=25)
        self.nombre_txt.grid(row=2, column=1, sticky="w", padx=10)
        self.nombre_txt.config(font=("Fixedsys",8,"bold"))

        # --------------APELLIDOS-----------------#
        self.lb_apellidos = Label(self.Frame_reg, text="Apellidos:")
        self.lb_apellidos.grid(row=3, column=0, padx=10, pady=10)
        self.lb_apellidos.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.apellidos_txt = Entry(self.Frame_reg,width=25)
        self.apellidos_txt.grid(row=3, column=1, sticky="w", padx=10)
        self.apellidos_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------EDAD-----------------#
        self.lb_edad = Label(self.Frame_reg, text="Edad:")
        self.lb_edad.grid(row=4, column=0, padx=10, pady=10)
        self.lb_edad.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.edad_txt = Entry(self.Frame_reg,width=25)
        self.edad_txt.grid(row=4, column=1, sticky="w", padx=10)
        self.edad_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------CARRERA-----------------#
        self.lb_carrera = Label(self.Frame_reg, text="Carrera:")
        self.lb_carrera.grid(row=5, column=0, padx=10, pady=10)
        self.lb_carrera.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.carrera_txt = Combobox(self.Frame_reg, width=23, state="readonly")
        self.carrera_txt.grid(row=5, column=1, sticky="w", padx=10)
        self.carrera_txt.config(font=("Fixedsys", 8, "bold"))
        self.carrera_txt["values"] = ["Derecho"]
        self.carrera_txt.bind("<<ComboboxSelected>>", self.opcionCarrera)

        # --------------SEMESTRE-----------------#
        self.lb_semestre = Label(self.Frame_reg, text="Semestre:")
        self.lb_semestre.grid(row=6, column=0, padx=10, pady=10)
        self.lb_semestre.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.semestre_txt = Combobox(self.Frame_reg, width=23, state="readonly")
        self.semestre_txt.grid(row=6, column=1, sticky="w", padx=10)
        self.semestre_txt.config(font=("Fixedsys", 8, "bold"))
        self.semestre_txt.bind("<<ComboboxSelected>>", self.opcionSemestre)

        # --------------Nivel-----------------#
        self.lb_nivel = Label(self.Frame_reg, text="Nivel:")
        self.lb_nivel.grid(row=7, column=0, padx=10, pady=10)
        self.lb_nivel.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.nivel_txt = Combobox(self.Frame_reg, width=23, state="readonly")
        self.nivel_txt.grid(row=7, column=1, sticky="w", padx=10)
        self.nivel_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------SELECCION-----------------#
        self.lb_seleccion = Label(self.Frame_reg, text="Seleccion:")
        self.lb_seleccion.grid(row=8, column=0, padx=10, pady=10)
        self.lb_seleccion.config(font=("Fixedsys", 12, "bold"), fg="blue")
        self.seleccion_txt = Combobox(self.Frame_reg,width=23,state="readonly")
        self.seleccion_txt.grid(row=8, column=1, sticky="w", padx=10)
        self.seleccion_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------PROMEDIO-----------------#
#        self.lb_promedio = Label(self.Frame_reg, text="Promedio:")
#       self.lb_promedio.grid(row=9, column=0, padx=10, pady=10)
#        self.lb_promedio.config(font=("Fixedsys", 12, "bold"), fg="blue")
#        self.promedio_txt = Entry(self.Frame_reg,width=25)
#        self.promedio_txt.grid(row=9, column=1, sticky="w", padx=10)
#        self.promedio_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------USUARIO-----------------#
#        self.lb_user = Label(self.Frame_reg, text="Usuario:")
#        self.lb_user.grid(row=7, column=0, padx=10, pady=10)
#        self.lb_user.config(font=("Fixedsys", 12, "bold"), fg="blue")
#        self.usuario_txt = Entry(self.Frame_reg,width=25)
#        self.usuario_txt.grid(row=7, column=1, sticky="w", padx=10)
#        self.usuario_txt.config(font=("Fixedsys", 8, "bold"))

        # --------------CONTRASEÑA-----------------#
#        self.lb_pass = Label(self.Frame_reg, text="Password:")
#        self.lb_pass.grid(row=8, column=0, padx=10, pady=10)
#        self.lb_pass.config(font=("Fixedsys", 12, "bold"), fg="blue")
#        self.user_pass_txt = Entry(self.Frame_reg,width=25)
#        self.user_pass_txt.grid(row=8, column=1, sticky="w", padx=10)
#        self.user_pass_txt.config(show="*",font=("Fixedsys", 8, "bold"))

        # ---------------BOTONES-------------------#
                        # GUARDAR#
        self.btn_registro = Button(self.Frame_reg, text="Guardar", width=24,command=self.validar)
        self.btn_registro.grid(row=10, column=0, padx=10, pady=20, columnspan=2)
        self.btn_registro.config(bg="gray", fg="white",font=("Fixedsys", 12, "bold"))


class Anotado(Frame):

    def __init__(self,bd,usu,id_usu,master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Sistema de Anotado")
        self.master.resizable(False,False)
        self.pack()
        self.base = bd
        self.usuario = usu
        self.id_usuario = id_usu
        self.ruta_semestre = ""
        self.elemento = ""
        self.crear_widget()
        self.master.protocol("WM_DELETE_WINDOW", self.master.quit)

    def directorio(self):
        return os.listdir(self.ruta)

    def textos(self,event):
        self.cuadro_texto.config(state="normal")
        self.cuadro_texto.delete("1.0", "end")
        self.ruta_semestre=self.ruta
        #self.list_txt_bd = self.base.lista_textos_base(self.usuario)
        self.lista_texto = os.listdir(self.ruta_semestre)
        self.lista_texts = []
        self.lista_textos = []
        for lis in self.lista_texto:
            self.lista_texts.append(lis)
        if(self.list_archivos.size()<=0):
            self.list_archivos.insert(0,*self.lista_texts)
        else:
            self.list_archivos.delete(0,END)
            self.list_archivos.insert(0,*self.lista_texts)

    def seleccion(self):
        if (self.palabra_compleja_txt.get() == ""):
            messagebox.showwarning("¡Atención!", "Ingrese una palabra")
        else:
            palabra = self.palabra_compleja_txt.get()
            palabra_len = len(palabra)
            pal_busc_split = palabra.split()
            cont = 0
            self.repeticiones = 0
            self.pos_pal_bd = 0
            palabra_original = palabra
            index = '1.0'
            while True:
                plbra_i = self.cuadro_texto.search(palabra, index, stopindex=END)
                if not plbra_i:
                    break
                plbra_i_ini = int(plbra_i.split('.')[0])
                plbra_i_fin = int(plbra_i.split('.')[1]) + palabra_len
                plbra_fin_aux = int(plbra_i_fin + 1)
                plbra_ini_aux = int(plbra_i.split('.')[1])-1
                coords = '{}.{}'.format(plbra_i_ini, plbra_i_fin)
                aux_fin = self.cuadro_texto.get(str(plbra_i_ini) + "." + str(plbra_i_fin),
                                            str(plbra_i_ini) + "." + str(plbra_fin_aux))
                aux_ini = self.cuadro_texto.get(str(plbra_i_ini) + "." + str(plbra_ini_aux),
                                               str(plbra_i_ini) + "." + str(plbra_i.split('.')[1]))
                if(plbra_ini_aux==-1):
                    if (aux_fin == " " or aux_fin == "," or aux_fin == "." or aux_fin == ";" or aux_fin == ":"
                            or aux_fin == "?" or aux_fin == ")" or aux_fin == "]" or aux_fin == "}" or aux_fin == "|"):
                        self.cuadro_texto.tag_add("palabra compleja", plbra_i, coords)
                        self.cuadro_texto.tag_configure("palabra compleja", background="green", foreground="white",
                                                        font='helvetica 11 bold')
                        self.repeticiones += 1
                        if(cont==0):
                            self.pal_bd = palabra_original
                            self.long_pal_bd = len(self.pal_bd)
                            self.ced_bd = self.base.Obtener_Cedula(self.usuario)
                            for ced in self.ced_bd:
                                self.ced_usu_bd = ced["cedula"]
                            self.nom_usu_bd = self.base.Obt_nombres(self.usuario)
                            self.ape_usu_bd = self.base.Obt_apellidos(self.usuario)
                            self.texto_bd = self.list_archivos.get(self.indice)
                            self.cont_sep_sin_puntos = self.contenido
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(".", " ").replace(",", " ").replace(":", " ").replace(";", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(",", " ").replace("[", " ").replace("]", " ").replace("(", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(")", " ").replace("{", " ").replace("}", " ").replace("<", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(">", " ").replace("¿", " ").replace("?", " ").replace("!", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace("¡", " ")
                            self.contenido_bd = self.cont_sep_sin_puntos.split()
                            cont_palb_conj = 0
                            repeticion = 0
                            for i in range(len(self.contenido_bd)):
                                cont_palb_conj = 0
                                if (self.contenido_bd[i] == pal_busc_split[cont_palb_conj]):
                                    cont_palb_conj += 1
                                    k = i + 1
                                    if (repeticion == 0):
                                        self.pos_pal_bd = i + 1
                                    if (len(pal_busc_split) > 1):
                                        for l in range(k, (k + (len(pal_busc_split) - 1))):
                                            if (self.contenido_bd[l] == pal_busc_split[cont_palb_conj]):
                                                cont_palb_conj += 1
                                                i = l
                                            else:
                                                cont_palb_conj = 0
                                                break
                                        if (cont_palb_conj == len(pal_busc_split)):
                                            repeticion += 1
                                    else:
                                        repeticion += 1
                            cont += 1
                            index = coords
                        else:
                            cont += 1
                            index = coords
                    else:
                        index = coords
                else:
                    if ((aux_fin == " " or aux_fin == "," or aux_fin == "." or aux_fin == ";" or aux_fin==":"
                         or aux_fin == "?" or aux_fin == ")" or aux_fin == "]" or aux_fin == "}" or aux_fin == "|")
                        and
                        (aux_ini == " " or aux_ini == "," or aux_ini == "." or aux_ini == ";" or aux_ini==":"
                         or aux_ini == "¿" or aux_ini == "(") or aux_ini == "[" or aux_ini == "{" or aux_ini == "|"):
                        self.cuadro_texto.tag_add("palabra compleja", plbra_i, coords)
                        self.cuadro_texto.tag_configure("palabra compleja", background="green", foreground="white",
                                                    font='helvetica 11 bold')
                        self.repeticiones+=1
                        if(cont==0):
                            self.pal_bd = palabra_original
                            self.long_pal_bd = len(self.pal_bd)
                            self.ced_bd = self.base.Obtener_Cedula(self.usuario)
                            for ced in self.ced_bd:
                                self.ced_usu_bd = ced["cedula"]
                            self.nom_usu_bd = self.base.Obt_nombres(self.usuario)
                            self.ape_usu_bd = self.base.Obt_apellidos(self.usuario)
                            self.texto_bd = self.list_archivos.get(self.indice)
                            self.cont_sep_sin_puntos = self.contenido
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(".", " ").replace(",", " ").replace(":", " ").replace(";", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(",", " ").replace("[", " ").replace("]", " ").replace("(", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(")", " ").replace("{", " ").replace("}", " ").replace("<", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace(">", " ").replace("¿", " ").replace("?", " ").replace("!", " ")
                            self.cont_sep_sin_puntos = self.cont_sep_sin_puntos.replace("¡", " ")
                            self.contenido_bd = self.cont_sep_sin_puntos.split()
                            cont_palb_conj = 0
                            repeticion = 0
                            for i in range(len(self.contenido_bd)):
                                cont_palb_conj = 0
                                if (self.contenido_bd[i] == pal_busc_split[cont_palb_conj]):
                                    cont_palb_conj += 1
                                    k = i + 1
                                    if (repeticion == 0):
                                        self.pos_pal_bd = i + 1
                                    if (len(pal_busc_split) > 1):
                                        for l in range(k, (k + (len(pal_busc_split) - 1))):
                                            if (self.contenido_bd[l] == pal_busc_split[cont_palb_conj]):
                                                cont_palb_conj += 1
                                                i = l
                                            else:
                                                cont_palb_conj = 0
                                                break
                                        if (cont_palb_conj == len(pal_busc_split)):
                                            repeticion += 1
                                    else:
                                        repeticion += 1
                            cont += 1
                            index = coords
                        else:
                            cont += 1
                            index = coords
                    else:
                        index = coords
            if (cont < 1):
                messagebox.showwarning("¡Atención!", "Palabra no encontrada")
            else:
                data = {
                    'Palabra': self.pal_bd,
                    'Cedula': self.ced_usu_bd,
                    'Nombres': self.nom_usu_bd,
                    'Apellidos': self.ape_usu_bd,
                    'Nombre_archivo': self.texto_bd,
                    'Posicion_en_el_texto': self.pos_pal_bd,
                    'Longitud': self.long_pal_bd,
                    'Repeticiones': self.repeticiones
                }
                consult_selec = self.base.Consultar_seleccion(data)
                if(len(consult_selec ) > 0):
                    messagebox.showwarning("¡Atención!", "La palabra ya ha sido seleccionada en este texto")
                else:
                    self.base.Guardar_seleccion(data)

    def mostrar_texto(self, event):
        self.cuadro_texto.config(state="normal")
        self.cuadro_texto.delete("1.0", "end")
        try:
            self.indice = int(self.list_archivos.curselection()[0])
        except IndexError:
            e=False
        if(self.indice == -1):
            messagebox.showinfo("¡En hora buena!", "Textos terminados")
        else:
            self.elemento = self.list_archivos.get(self.indice)
            if(self.elemento!=""):
                self.ruta_text = self.ruta_semestre + self.elemento
                try:
                    self.archi = open(self.ruta_text,"r",encoding='utf-8')
                    self.contenido = self.archi.read()
                except UnicodeDecodeError:
                    self.archi = open(self.ruta_text,"r",encoding='latin-1')
                    self.contenido = self.archi.read()
                self.cuadro_texto.insert(1.0, self.contenido)
                self.cuadro_texto.config(state="disable")
                self.archi.close()
            else:
                self.cuadro_texto.config(state="disable")
                messagebox.showinfo("¡Terminado!", "Textos terminados en esta semestre")

    def culminar(self):
        if(self.elemento!=""):
            self.finalizar=messagebox.askokcancel("Finalizar", "¿Desea terminar el proceso de anotado en este texto?")
            if(self.finalizar== True):
                #self.base.Fin_anotado_text(self.usuario,self.list_archivos.get(self.indice))
                self.archivoTexto = self.ruta + self.list_archivos.get(self.indice)
                #self.list_txt_bd = self.base.lista_textos_base(self.usuario)
                os.remove(self.archivoTexto)
                self.lista_texto = os.listdir(self.ruta)
                self.lista_texts = []
                for lis in self.lista_texto:
                    self.lista_texts.append(lis)
                if(self.list_archivos.size()<=0):
                    self.list_archivos.insert(0,*self.lista_texts)
                else:
                    self.list_archivos.delete(0,END)
                    self.list_archivos.insert(0,*self.lista_texts)
        else:
            messagebox.showinfo("¡Terminado!", "No hay texto en pantalla")

    def crear_widget(self):
        # --------------------------------------------------------------------------------#
        # FRAME PARA EL LISTADO DE TEXTOS#
        self.Frame_listado = Frame(self)
        self.Frame_listado.config(width=200, height=500, cursor="hand2", borderwidth="2")
        self.Frame_listado.grid(row=0, column=0, rowspan=2)
        self.lb_semestre = Label(self.Frame_listado, text="Semestres:")
        self.lb_semestre.grid(row=0, column=0, pady=5)
        self.ruta = self.base.Obt_ruta_textos(self.usuario)
        self.list_directorios = self.directorio()
        self.semestre_txt = Combobox(self.Frame_listado, width=30, state="readonly")
        self.semestre_txt.grid(row=0, column=1, sticky="w", pady=5)
        self.semestre_txt["values"] = self.base.Obt_sem_niv_sel(self.usuario).strip().split('\t')
        self.semestre_txt.bind("<<ComboboxSelected>>", self.textos)
        self.list_archivos = Listbox(self.Frame_listado, relief="groove", borderwidth="2", selectmode=EXTENDED)
        self.list_archivos.grid(row=1, column=0, sticky="w", columnspan=2)
        self.barra = Scrollbar(self.Frame_listado,command=self.list_archivos.yview)
        self.barra.grid(row=1, column=0, columnspan=2, sticky="nse")
        self.list_archivos.config(yscrollcommand=self.barra.set,width=40, height=29)
        self.list_archivos.bind("<<ListboxSelect>>", self.mostrar_texto)

        # --------------------------------------------------------------------------------#
        #FRAME PARA EL ENCABEZADO DE PENDIENTE Y SESION#

        self.Frame_encabezado = Frame(self)
        self.Frame_encabezado.config(width=1400, height=35, cursor="hand2", relief="groove")
        self.Frame_encabezado.grid(row=0, column=1)
        self.btn_cerrar_sesion = Button(self.Frame_encabezado, text="Cerrar Sesión", width=23, command=self.master.quit)
        self.btn_cerrar_sesion.grid(row=0, column=0, columnspan=4, sticky="w")
        self.btn_cerrar_sesion.config(bg="gray", fg="white")

        # --------------------------------------------------------------------------------#
        # FRAME PARA EL ANOTADO DE PALABRAS#

        self.Frame_anotado = Frame(self)
        self.Frame_anotado.config(width=1800, height=465, cursor="hand2")
        self.Frame_anotado.grid(row=1, column=1)
        self.archivo = None
        self.indice = -1
        self.cuadro_texto = Text(self.Frame_anotado, width=70, height=27, relief="groove", borderwidth="2")
        self.cuadro_texto.grid(row=0, column=0, columnspan=4)
#        self.barrax = Scrollbar(self.Frame_anotado,orient="horizontal",command=self.cuadro_texto.xview)
#        self.barrax.grid(row=0, column=0, columnspan=5, sticky="ews")
        self.barray = Scrollbar(self.Frame_anotado,command=self.cuadro_texto.yview)
        self.barray.grid(row=0, column=1, columnspan=4, sticky="nse")
        self.cuadro_texto.config(yscrollcommand=self.barray.set)
        self.lb_palabra_compeja = Label(self.Frame_anotado, text="Palabra compleja:")
        self.lb_palabra_compeja.grid(row=1, column=0)
        self.palabra_compleja_txt = Entry(self.Frame_anotado, width=24, relief="groove", borderwidth="2")
        self.palabra_compleja_txt.grid(row=1, column=1, sticky="w")
        self.btn_seleccionar = Button(self.Frame_anotado, text="Guardar Palabra", width=24, command=self.seleccion)
        self.btn_seleccionar.grid(row=1, column=2)
        self.btn_seleccionar.config(bg="gray", fg="white")
        self.btn_culminar = Button(self.Frame_anotado, text="Culminar", width=24, command=self.culminar)
        self.btn_culminar.grid(row=1, column=3)
        self.btn_culminar.config(bg="gray", fg="white")


class switch:
	def __init__(self, variable, comparator=None, strict=False):
		self.variable = variable
		self.matched = False
		self.matching = False
		if comparator:
			self.comparator = comparator
		else:
			self.comparator = lambda x, y: x == y
		self.strict = strict

	def __enter__(self):
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		pass

	def case(self, expr, break_=False):
		if self.strict:
			if self.matched:
				return False
		if self.matching or self.comparator(self.variable, expr):
			if not break_:
				self.matching = True
			else:
				self.matched = True
				self.matching = False
			return True
		else:
			return False

	def default(self):
		return not self.matched and not self.matching


#### CÓDIGO DESARROLLADO POR ALVAREZ TORRES NATHALY GISSELL - QUISHPI LEMA CARLOS LUIS ###
### PARA LA OBTENCIÓN DEL TÍTULO COMO INGENIERO EN SISTEMAS COMPUTACIONALES ###

inst_login = Tk()
vent_login = Login(inst_login)
vent_login.mainloop()
